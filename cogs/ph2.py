import discord
from discord.ext import commands
import asyncio
import openpyxl as op

class Phoenix2(commands.Cog):
    def __init__(self, client):
        self.client = client


    # phoenix 2 info and commands
    @commands.command(name="ships")
    async def ships(self, ctx):
        embed1=discord.Embed(title="Phoenix 2 Ships", color=0xde1212)
        embed1.set_thumbnail(url="https://firigames.com/images/app_phoenix2/poster.jpg")
        embed1.add_field(name="Shinova", value="\u200b", inline=True)
        embed1.add_field(name="NC-271", value="\u200b", inline=True)
        embed1.add_field(name="Veil", value="\u200b", inline=True)
        embed1.add_field(name="Jn'dur", value="\u200b", inline=True)
        embed1.add_field(name="Yoth-Hola", value="\u200b", inline=True)
        embed1.add_field(name="Razor VI", value="\u200b", inline=True)
        embed1.add_field(name="Mist Star", value="\u200b", inline=True)
        embed1.add_field(name="Proxymar", value="\u200b", inline=True)
        embed1.add_field(name="Trireme", value="\u200b", inline=True)
        embed1.add_field(name="Phoenix", value="\u200b", inline=True)
        embed1.add_field(name="Czar ", value="\u200b", inline=True)
        embed1.add_field(name="Vuxine", value="\u200b", inline=True)
        embed1.add_field(name="Arietis ", value="\u200b", inline=True)
        embed1.add_field(name="Essin", value="\u200b", inline=True)
        embed1.add_field(name="502-Q8", value="\u200b", inline=True)
        embed1.add_field(name="Gladius", value="\u200b", inline=True)
        embed1.add_field(name="Valkyrie", value="\u200b", inline=True)
        embed1.add_field(name="Yigothu", value="\u200b", inline=True)
        embed1.add_field(name="Exarch", value="\u200b", inline=True)
        embed1.add_field(name="Njorun", value="\u200b", inline=True)
        embed1.add_field(name="NC-150", value="\u200b", inline=True)
        embed1.add_field(name="Warden", value="\u200b", inline=True)
        embed1.add_field(name="Krillou", value="\u200b", inline=True)
        embed1.add_field(name="Jericho", value="\u200b", inline=True)
        embed1.set_footer(text="Page 1")

        embed2=discord.Embed(title="Phoenix 2 Ships", color=0xde1212)
        embed2.set_thumbnail(url="https://firigames.com/images/app_phoenix2/poster.jpg")
        embed2.add_field(name="Oran", value="\u200b", inline=True)
        embed2.add_field(name="Kada Veni", value="\u200b", inline=True)
        embed2.add_field(name="Havoc", value="\u200b", inline=True)
        embed2.add_field(name="Ogon", value="\u200b", inline=True)
        embed2.add_field(name="Xaniea", value="\u200b", inline=True)
        embed2.add_field(name="Atlas", value="\u200b", inline=True)
        embed2.add_field(name="EX04239", value="\u200b", inline=True)
        embed2.add_field(name="Prime X", value="\u200b", inline=True)
        embed2.add_field(name="Sakura", value="\u200b", inline=True)
        embed2.add_field(name="Zephyr", value="\u200b", inline=True)
        embed2.add_field(name="Hime", value="\u200b", inline=True)
        embed2.add_field(name="Hunter", value="\u200b", inline=True)
        embed2.add_field(name="Lokie", value="\u200b", inline=True)
        embed2.add_field(name="Saber ", value="\u200b", inline=True)
        embed2.add_field(name="Proteus", value="\u200b", inline=True)
        embed2.add_field(name="Orion", value="\u200b", inline=True)
        embed2.add_field(name="Sonah", value="\u200b", inline=True)
        embed2.add_field(name="Antioch", value="\u200b", inline=True)
        embed2.add_field(name="Trinity", value="\u200b", inline=True)
        embed2.add_field(name="Shogun", value="\u200b", inline=True)
        embed2.add_field(name="Predator", value="\u200b", inline=True)
        embed2.add_field(name="Icarus", value="\u200b", inline=True)
        embed2.add_field(name="Disaris", value="\u200b", inline=True)
        embed2.add_field(name="Stinger", value="\u200b", inline=True)
        embed2.add_field(name="Esperon", value="\u200b", inline=True)
        embed2.set_footer(text="Page 2")

        embed3=discord.Embed(title="Phoenix 2 Ships", color=0xde1212)
        embed3.set_thumbnail(url="https://firigames.com/images/app_phoenix2/poster.jpg")
        embed3.add_field(name="Tar'cah", value="\u200b", inline=True)
        embed3.add_field(name="Fujin", value="\u200b", inline=True)
        embed3.add_field(name="Tempest", value="\u200b", inline=True)
        embed3.add_field(name="Wrackr", value="\u200b", inline=True)
        embed3.add_field(name="Cinnri", value="\u200b", inline=True)
        embed3.add_field(name="Buhloo", value="\u200b", inline=True)
        embed3.add_field(name="Geist", value="\u200b", inline=True)
        embed3.add_field(name="Tilla'tor", value="\u200b", inline=True)
        embed3.add_field(name="Elyon", value="\u200b", inline=True)
        embed3.add_field(name="Aurora", value="\u200b", inline=True)
        embed3.add_field(name="Gorthaur", value="\u200b", inline=True)
        embed3.add_field(name="Heechi", value="\u200b", inline=True)
        embed3.add_field(name="Vani-Vith", value="\u200b", inline=True)
        embed3.add_field(name="Baqlor", value="\u200b", inline=True)
        embed3.add_field(name="Dragonfly", value="\u200b", inline=True)
        embed3.add_field(name="Barret", value="\u200b", inline=True)
        embed3.add_field(name="Torrent", value="\u200b", inline=True)
        embed3.add_field(name="Boxer", value="\u200b", inline=True)
        embed3.add_field(name="Corsair", value="\u200b", inline=True)
        embed3.add_field(name="UHB", value="\u200b", inline=True)
        embed3.add_field(name="Wraith", value="\u200b", inline=True)
        embed3.add_field(name="Lorilou", value="\u200b", inline=True)
        embed3.add_field(name="AB8/KLYN", value="\u200b", inline=True)
        embed3.add_field(name="Qhelqod", value="\u200b", inline=True)
        embed3.add_field(name="Jeria", value="\u200b", inline=True)
        embed3.set_footer(text="Page 3")

        embed4=discord.Embed(title="Phoenix 2 Ships", color=0xde1212)
        embed4.set_thumbnail(url="https://firigames.com/images/app_phoenix2/poster.jpg")
        embed4.add_field(name="Nimbus", value="\u200b", inline=True)
        embed4.add_field(name="Monsoon", value="\u200b", inline=True)
        embed4.add_field(name="Marauder", value="\u200b", inline=True)
        embed4.add_field(name="Xavis", value="\u200b", inline=True)
        embed4.add_field(name="Starless", value="\u200b", inline=True)
        embed4.add_field(name="Zhetass", value="\u200b", inline=True)
        embed4.add_field(name="Zimitr", value="\u200b", inline=True)
        embed4.add_field(name="Kibarrax", value="\u200b", inline=True)
        embed4.add_field(name="Scuuxun", value="\u200b", inline=True)
        embed4.add_field(name="Juggernaut", value="\u200b", inline=True)
        embed4.add_field(name="Von Braun", value="\u200b", inline=True)
        embed4.add_field(name="X-81", value="\u200b", inline=True)
        embed4.add_field(name="Mirage", value="\u200b", inline=True)
        embed4.add_field(name="Lyova", value="\u200b", inline=True)
        embed4.add_field(name="Widget", value="\u200b", inline=True)
        embed4.add_field(name="Claymore", value="\u200b", inline=True)
        embed4.add_field(name="Reaper", value="\u200b", inline=True)
        embed4.add_field(name="Banshee", value="\u200b", inline=True)
        embed4.add_field(name="Centurion", value="\u200b", inline=True)
        embed4.add_field(name="Naya", value="\u200b", inline=True)
        embed4.add_field(name="Mistral", value="\u200b", inline=True)
        embed4.add_field(name="Pandora", value="\u200b", inline=True)
        embed4.add_field(name="Neni", value="\u200b", inline=True)
        embed4.set_footer(text="Page 4")


        shippages = [embed1, embed2, embed3, embed4]
        buttons = [u"\u23EA", u"\u25C0", u"\u25B6", u"\u23E9"]
        current = 0
        msg = await ctx.send(embed=shippages[current])

        for button in buttons:
            await msg.add_reaction(button)
        while True:
            try:
                reaction, user = await self.client.wait_for("reaction_add", check=lambda reaction, user: user == ctx.author and not user.bot and reaction.message == msg and reaction.emoji in buttons, timeout=60.0)
            except asyncio.TimeoutError:
                embed = shippages[current]
                embed.set_footer(text="Timed Out")
                await msg.clear_reactions()

            else:
                previous_page = current
                if reaction.emoji == "\u23EA":
                    current = 0
                elif reaction.emoji == "\u25C0":
                    if current > 0:
                        current -= 1
                
                elif reaction.emoji == "\u25B6":
                    if current < len(shippages)-1:
                        current += 1
                
                elif reaction.emoji == "\u23E9":
                    current = len(shippages)-1

                await msg.remove_reaction(reaction.emoji, ctx.author)

                if current != previous_page:
                    await msg.edit(embed=shippages[current])


    @commands.command(name="ship")
    async def ship(self, ctx, shipname: str):
        wb = op.load_workbook('funnies.xlsx')
        sheet = wb['Sheet1']
        headers = [cell.value for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column) for cell in row][:sheet.max_column]
        name_col = headers.index('name')
        for row in sheet.iter_rows(min_row=2):
            if row[name_col].value == shipname:
                name = row[name_col].value
                mainType = row[headers.index('mainType')].value
                aura = row[headers.index('aura')].value
                zen = row[headers.index('zen')].value
                apex1 = row[headers.index('apex1')].value
                apex2 = row[headers.index('apex2')].value
                jimNote = row[headers.index('jimNote')].value
                img = row[headers.index('img')].value
                # create an embed or return a message containing the information
                embed = discord.Embed(title=name, description=mainType, color=0xba0d0d)
                embed.set_thumbnail(url=img)
                embed.add_field(name="Aura", value=aura, inline=True)
                embed.add_field(name="Zen", value=zen, inline=True)
                embed.add_field(name="Apex 1", value=apex1, inline=False)
                embed.add_field(name="Apex 2", value=apex2, inline=False)
                embed.set_footer(text=jimNote)
                await ctx.send(embed=embed)
                return
        await ctx.send(f"Could not find {shipname} in the excel sheet")

async def setup(client):
    await client.add_cog(Phoenix2(client))